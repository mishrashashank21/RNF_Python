import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from datetime import time, datetime, timedelta
import time
import os.path
from openpyxl import load_workbook

begin_time = datetime.now()
print("Begin time: {}".format(begin_time))
period = datetime(2021,12,18,00,21)
op = webdriver.ChromeOptions()
p = {'download.default_directory': r'C:\Users\Dell\Downloads\Session_history\android'}
# p = {'download.default_directory': r'C:\Users\RNF-Administrator.DESKTOP-1HQ96LI\Documents\Session_History'}
# q = r'C:\Users\RNF-Administrator.DESKTOP-1HQ96LI\Documents\Session_History'
q = r'C:\Users\Dell\Downloads\Session_history\android'
op.add_experimental_option('prefs',p)
# while not os.path.exists(q):
#     time.sleep(1)
# driver = webdriver.Chrome(executable_path = r"C:\Users\RNF-Administrator.DESKTOP-1HQ96LI\OneDrive\Python Scripts\PyCharm\chromedriver_win32\chromedriver.exe", options=op)

driver = webdriver.Chrome(executable_path = r"C:\Users\Dell\OneDrive\Python Scripts\PyCharm\chromedriver_win32\chromedriver.exe", options=op)

driver.get("https://7starslots.com:8042/login#/index")
# driver.get("https://7starslots.com:8042/dashboard#/buy er_session_data?user_id=7HCIOS26FKR3YW6ZM")
# driver.maximize_window()

username = driver.find_element_by_name("email")
password = driver.find_element_by_name("password")

username.send_keys("shashank.mishra@phonato.com")
password.send_keys("shashank#321")

driver.find_element_by_xpath("//*[@id='login-form']/div[3]/div/button").click()
driver.find_element_by_xpath("/html/body/div[2]/aside[1]/section/ul/li[1]/a/span/select/option[2]").click()

def session_data():
    driver.find_element_by_xpath("/html/body/div[2]/aside[1]/section/ul/li[6]/a/span[1]").click()
    time.sleep(3)
    driver.find_element_by_xpath("/html/body/div[2]/aside[1]/section/ul/li[6]/ul/li[7]/a/span").click()
    time.sleep(3)

session_data()

# fromdate_list = []
# fromdate_list.extend([int(x) for x in input("Input from date(year,month,date):" ).split(",")])
# fromdate = datetime(int(fromdate_list[0]),int(fromdate_list[1]),int(fromdate_list[2]))
# print(fromdate.date())
# print(fromdate.day)
# print(fromdate.month)
# fromdate_day = fromdate.day
# fromdate_month = fromdate.month
#
# todate_list = []
# todate_list.extend([int(x) for x in input("Input to date(year,month,date):" ).split(",")])
# todate = datetime(int(todate_list[0]),int(todate_list[1]),int(todate_list[2]))
# print(todate.date())
# print(todate.day)
# print(todate.month)
# todate_day = todate.day
# todate_month = todate.month

# driver.find_element_by_id("from_date").click()
# alldates = driver.find_elements_by_xpath("//td[@class='day' or @class='today day']")
# for dateelement in alldates:
#     date = int(dateelement.text)
#     print(date)
#     if date == fromdate.day:
#         dateelement.click()
#         break
#
# driver.find_element_by_id("to_date").click()
# alldates = driver.find_elements_by_xpath("//td[@class='day' or @class='today day']")
# for dateelement in alldates:
#     date = int(dateelement.text)
#     print(date)
#     if date == todate.day:
#         dateelement.click()
#         break

# df = pd.DataFrame()
di = pd.read_csv('Dataset/7hc_andorid_philippines_Nov-10-Dec-10.csv')
# di = pd.read_excel('Dataset/7hc_andorid_buyers_Nov_1-15.xlsx',sheet_name="7hc_andorid_buyers_Nov_1-15")
# print(di.loc[0,'_id'])
# print(di.columns)
# print(di.shape)

# userid = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[2]/div/input')
# userid.send_keys('7HCAD0FKRWZ5PZM')
# driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[3]/button[1]').click()
# # driver.implicitly_wait(10)
# time.sleep(5)
# driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/button[1]').click()
# time.sleep(5)
# driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[3]/button[2]').click()
# reject_list = ["7HCAD154P3BM7HT8P"]
for i in range(len(di['_id'])):
    # if di.iloc[i,di.columns.get_loc('_id')] in reject_list:
    #     continue
    time_start = time.time()
    # count = 0
    # while time.time() <= time_start + 60:
    # if count > 0:
    #     break
    print("{}. {}".format(i+1,di.iloc[i, di.columns.get_loc('_id')]))
    fromdate_sec = di.iloc[i, di.columns.get_loc('_created_at')]
    todate_sec = di.iloc[i, di.columns.get_loc('_updated_at')]
    conversion_from = timedelta(seconds=int(fromdate_sec))
    fromdate_final = (period - conversion_from)
    print(fromdate_final)
    conversion_to = timedelta(seconds=int(todate_sec))
    todate_final = (period - conversion_to)
    todate_final_new = todate_final
    print(todate_final)
    # print(todate_final.year)
    # print(todate_final.month)
    # print(todate_final.day)
    # time.sleep(2)
    # driver.find_element_by_id("from_date").click()
    driver.find_element_by_xpath("/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[1]/div[1]/input").click()
    # if i ==0:
    #     driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()

    driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
    fromdate_year = int(driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").text)

    loop =  int(fromdate_final.year) -fromdate_year
    print("from_loop: {}".format(loop))
    if loop < 0:
        for click in range(abs(loop)):
            time.sleep(1)
            driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[1]").click()

    if loop > 0:
        for click in range(abs(loop)):
            driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[3]").click()

    # allyears_from = driver.find_elements_by_xpath("//td/span[@class='year old' or @class='year focused' or @class='year new' or @class='year' or @class='year active focused']")
    # for yearelement_from in allyears_from:
    #     year_from = int(yearelement_from.text)
    #     if year_from == fromdate_final.year:
    #         yearelement_from.click()
    #         break

    allmonths_from = driver.find_elements_by_xpath("//td/span[@class='month' or @class='month focused' or @class='month focused active']")
    for monthelement_from in allmonths_from:
        month_from = monthelement_from.text
        if month_from == fromdate_final.strftime('%b'):
            monthelement_from.click()
            break

    alldates_from = driver.find_elements_by_xpath("//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
    # time.sleep(1)
    for dateelement_from in alldates_from:
        date_from = int(dateelement_from.text)
        # print(date_from)
        if date_from == fromdate_final.day:
            dateelement_from.click()
            break

    driver.find_element_by_id("to_date").click()
    # if i ==0:
    #     driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()
    # time.sleep(1)

    driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
    # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
    todate_year = int(driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").text)
    loop = int(todate_final.year) - todate_year
    print("to_loop: {}".format(loop))
    if loop < 0:
        for click in range(abs(loop)):
            driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[1]").click()

    if loop > 0:
        for click in range(abs(loop)):
            driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[3]").click()


    # allyears_to = driver.find_elements_by_xpath( "//td/span[@class='year old' or @class='year focused' or @class='year new' or @class='year' or @class='year active focused']")
    # for yearelement_to in allyears_to:
    #     year_to = int(yearelement_to.text)
    #     if year_to == todate_final.year:
    #         yearelement_to.click()
    #         break

    allmonths_to = driver.find_elements_by_xpath("//td/span[@class='month' or @class='month focused' or @class='month focused active']")
    for monthelement_to in allmonths_to:
        month_to = monthelement_to.text
        if month_to == todate_final.strftime('%b'):
            monthelement_to.click()
            break

    alldates_to = driver.find_elements_by_xpath("//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
    for dateelement_to in alldates_to:
        date_to = int(dateelement_to.text)
        # print(date_to)

        if date_to == todate_final.day:
            dateelement_to.click()
            break

    userid = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[2]/div/input')
    # userid.sendkeys(di.loc[i,'_id'])
    userid.send_keys(di.iloc[i,:1].item())
    # print(di.iloc[i,:1].item())
    search = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[3]/button[1]')

    # time_start = time.time()
    while not search.is_enabled():
        time.sleep(1)
        # if time.time() > time_start + 10:
        #     break

    if search.is_enabled():
        search.click()
        # time_start = time.time()
        while not search.is_enabled():
            time.sleep(1)
            # if time.time() > time_start + 60:
            #     session_data()
            #     break
    #             driver.get("https://7starslots.com:8042/dashboard#/buyer_session_data")
    #             # driver.refresh()
    #             # time.sleep(5)
    #             break
    # if time.time() > time_start + 60:
    #     continue

    expo = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/button[1]')
    forward = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/button[3]')

    # time_start0 = time.time()
    count = 0
    while not expo.is_enabled():
        search = driver.find_element_by_xpath(
            '/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[3]/button[1]')

        expo = driver.find_element_by_xpath(
            '/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/button[1]')
        forward = driver.find_element_by_xpath(
            '/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/button[3]')
        # c = driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[2]/div/div[1]/div/label/label[3]')
        # e = int(c.text)
        # if e > 5000:
        #     break
        # print(e)
        if forward.is_enabled():
            driver.find_element_by_id("to_date").click()
            # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()
            # time.sleep(1)
            alldates_to = driver.find_elements_by_xpath("//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
            for dateelement_to in alldates_to:
                date_to = int(dateelement_to.text)
                # print(date_to)
                if date_to == todate_final.day:
                    dateelement_to.click()
                    break

            todate_final = todate_final_new - timedelta(days=count)
            print(todate_final)
            driver.find_element_by_id("to_date").click()
            driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
            # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
            todate_year = int(driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").text)
            loop = int(todate_final.year) - todate_year
            print("to_loop: {}".format(loop))
            if loop < 0:
                for click in range(abs(loop)):
                    driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[1]").click()

            if loop > 0:
                for click in range(abs(loop)):
                    driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[3]").click()

            allmonths_to = driver.find_elements_by_xpath(
                "//td/span[@class='month' or @class='month focused' or @class='month focused active']")
            for monthelement_to in allmonths_to:
                month_to = monthelement_to.text
                if month_to == todate_final.strftime('%b'):
                    monthelement_to.click()
                    break

            alldates_to = driver.find_elements_by_xpath(
                "//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
            for dateelement_to in alldates_to:
                date_to = int(dateelement_to.text)
                # print(date_to)

                if date_to == todate_final.day:
                    dateelement_to.click()
                    break

            # time_start = time.time()
            while not search.is_enabled():
                time.sleep(1)
                # if time.time() > time_start + 10:
                #     break

            if search.is_enabled():
                search.click()
                # time_start = time.time()
                while not search.is_enabled():
                    time.sleep(1)
                    # if time.time() > time_start + 60:
                    #     session_data()
                    #     break
                    # if time.time() > time_start + 60:
                    #     driver.get("https://7starslots.com:8042/dashboard#/buyer_session_data")
                    #     # time.sleep(5)
                    #     # driver.refresh()
                    #     break
            # if time.time() > time_start + 60:
            #     break

            # time_start1 = time.time()
            count = count + 15
            # while not expo.is_enabled():
            #     if forward.is_enabled():
            #         driver.find_element_by_id("to_date").click()
            #         # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()
            #         # time.sleep(1)
            #         alldates_to = driver.find_elements_by_xpath(
            #             "//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
            #         for dateelement_to in alldates_to:
            #             date_to = int(dateelement_to.text)
            #             # print(date_to)
            #             if date_to == todate_final.day:
            #                 dateelement_to.click()
            #                 break
            #         todate_final = fromdate_final + timedelta(days=7)
            #         driver.find_element_by_id("to_date").click()
            #         driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
            #         # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
            #         todate_year = int(
            #             driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").text)
            #         loop = int(todate_final.year) - todate_year
            #         print("to_loop: {}".format(loop))
            #         if loop < 0:
            #             for click in range(abs(loop)):
            #                 driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[1]").click()
            #
            #         if loop > 0:
            #             for click in range(abs(loop)):
            #                 driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[3]").click()
            #
            #         allmonths_to = driver.find_elements_by_xpath(
            #             "//td/span[@class='month' or @class='month focused' or @class='month focused active']")
            #         for monthelement_to in allmonths_to:
            #             month_to = monthelement_to.text
            #             if month_to == todate_final.strftime('%b'):
            #                 monthelement_to.click()
            #                 break
            #
            #         alldates_to = driver.find_elements_by_xpath(
            #             "//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
            #         for dateelement_to in alldates_to:
            #             date_to = int(dateelement_to.text)
            #             # print(date_to)
            #
            #             if date_to == todate_final.day:
            #                 dateelement_to.click()
            #                 break
            #
            #         # time_start = time.time()
            #         while not search.is_enabled():
            #             time.sleep(1)
            #             # if time.time() > time_start + 10:
            #             #     break
            #
            #         if search.is_enabled():
            #             search.click()
            #             # time_start = time.time()
            #             while not search.is_enabled():
            #                 time.sleep(1)
            #                 # if time.time() > time_start + 60:
            #                 #     session_data()
            #                 #     break
            #     #             if time.time() > time_start + 60:
            #     #                 driver.get("https://7starslots.com:8042/dashboard#/buyer_session_data")
            #     #                 # time.sleep(5)
            #     #                 # driver.refresh()
            #     #                 break
            #     #
            #     # if time.time() > time_start + 60:
            #     #     break
            #
            #     elif search.is_enabled() and not expo.is_enabled() and not forward.is_enabled():
            #         print("No data found inner loop")
            #         break
                # if time.time() > time_start1 + 10:
                #     break
                # time.sleep(1)
            # break
        elif search.is_enabled() and not expo.is_enabled() and not forward.is_enabled():
            print("No data found")
            break
        # if time.time() > time_start0 + 10:
        #     break

        # if time.time() > time_start + 60:
        #     break
        time.sleep(1)

    filepath = q + '\{}.csv'.format(di.loc[i, '_id'])
    if expo.is_enabled():
        expo.click()
        # filepath = r'C:\Users\Dell\Downloads\Session_history\{}.csv'.format(di.loc[i,'_id'])

        # time_start = time.time()
        while not os.path.isfile(filepath):
            #     break
            # if time.time() > time_start + 10:
            #     break
            time.sleep(1)
        # if not expo.is_enabled() and not forward.is_enabled():
        #     break
    # if os.path.isfile(filepath) == False:
    #     # df = df.append(di.iloc[i])
    #     df = pd.DataFrame(di.iloc[i])
    #     # df.to_excel('exports/Session_history_not_downloaded_check.xlsx', index=False)
    #     book = load_workbook('exports/Session_history_not_downloaded.xlsx')
    #     writer = pd.ExcelWriter('exports/Session_history_not_downloaded.xlsx', engine='openpyxl')
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     # df.to_excel('exports/Session_history_not_downloaded_check.xlsx', index=False)
    #     df.to_excel(writer, startrow=writer.sheets[book.worksheets[0].title].max_row, startcol=0, header=False,
    #                 index=False)
    #     writer.save()
    #     writer.close()

    driver.find_element_by_xpath('/html/body/div[2]/div[1]/div/div/section[2]/div/div/div[1]/div/form/div[3]/button[2]').click()
    id = userid.get_attribute("value")
    while not id == '':
        time.sleep(1)


    driver.find_element_by_id("from_date").click()
    # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()


    alldates_from = driver.find_elements_by_xpath("//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
    # time.sleep(1)
    for dateelement_from in alldates_from:
        date_from = int(dateelement_from.text)
        # print(date_from)
        if date_from == fromdate_final.day:
            dateelement_from.click()
            break

    # if fromdate_final.year < 2019:
    #     driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
    #     driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").text
    #     driver.find_element_by_xpath("/html/body/div[3]/div[3]/table/thead/tr[2]/th[3]").click()

    driver.find_element_by_id("to_date").click()
    # driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[1]").click()
    # time.sleep(1)
    alldates_to = driver.find_elements_by_xpath("//td[@class='day' or @class='today day' or @class='active day' or @class='today active day']")
    for dateelement_to in alldates_to:
        date_to = int(dateelement_to.text)
        # print(date_to)
        if date_to == todate_final.day:
            dateelement_to.click()
            break
    # count = count + 1

# if time.time() >= time_start + 60:
#     driver.find_element_by_xpath("/html/body/div[2]/aside[1]/section/ul/li[2]/a/span").click()
#     time.sleep(3)
#     driver.find_element_by_xpath("/html/body/div[2]/aside[1]/section/ul/li[6]/ul/li[7]/a/span").click()
#     time.sleep(3)


    # if fromdate_final.year < 2019:
    #     driver.find_element_by_xpath("/html/body/div[3]/div[1]/table/thead/tr[2]/th[2]").click()
    #     driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/thead/tr[2]/th[2]").click()
    #     driver.find_element_by_xpath("/html/body/div[3]/div[3]/table/thead/tr[2]/th[3]").click()


driver.close()
print("Execution time: {}".format(datetime.now()-begin_time))
