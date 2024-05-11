import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import workbook
from datetime import datetime
from datetime import timedelta
import time
import os
import math

#Creating export excel sheet
begin_time = datetime.now()
print(begin_time)
path_export = r'C:\Users\Dell\Documents\Python Scripts\7hc_andorid_philippines_Nov-10-Dec-10_user-accountability_D137.xlsx'
if os.path.exists(path_export):
    os.remove(path_export)
    print("File removal successfull")
else:
    print("The file does not exist")


def df_new():
    df = pd.DataFrame(columns=['created_at_ist', '_id', 'Daily_Total_Bet_Placed', 'Daily_Max_Bet_Placed',
                                            'Daily_Games_Played', 'Daily_Games_Win', 'Daily_Games_Loose',
                                            'Daily_Level_Change', 'Daily_Level_Up_Amount', 'Daily_Win_Amount',
                                            'Daily_$_Spent', 300, 140, 97, 92, 85, 'EOD_Level_Up_Amount',
                                            'EOD_Win_Amount', 'EOD_Coins', 'EOD_Games_Played', 'EOD_Games_Win',
                                            'EOD_Games_Loose', 'EOD_Level', 'EOD_Total_Bet_Placed',
                                            'Number_of_Days_Played', 'Duration_of_Retention','EOD_$_Spent', '300.0', '140.0', '97.0', '92.0',
                                            '85.0','total_games_played', 'dollar_spent'])

    df.loc[0,['300.0', '140.0', '97.0', '92.0', '85.0']] = 0
    return df

df = pd.DataFrame(columns = ['created_at_ist','_id','Daily_Total_Bet_Placed','Daily_Max_Bet_Placed','Daily_Games_Played','Daily_Games_Win','Daily_Games_Loose','Daily_Level_Change','Daily_Level_Up_Amount','Daily_Win_Amount','Daily_$_Spent',300,140,97,92,85,'EOD_Level_Up_Amount','EOD_Win_Amount','EOD_Coins','EOD_Games_Played','EOD_Games_Win','EOD_Games_Loose','EOD_Level','EOD_Total_Bet_Placed','Number_of_Days_Played','Duration_of_Retention','EOD_$_Spent','300.0','140.0','97.0','92.0','85.0','total_games_played','dollar_spent'])
# df = pd.DataFrame(columns = ['created_at_ist','_id','Daily_Total_Bet_Placed','Daily_Max_Bet_Placed','Daily_Games_Played','Daily_Games_Win','Daily_Games_Loose','Daily_Level_Change','Daily_Level_Up_Amount','Daily_Win_Amount','Daily_$_Spent',300,140,97,92,85,'EOD_Level_Up_Amount','EOD_Win_Amount','EOD_Coins','EOD_Games_Played','EOD_Games_Win','EOD_Games_Loose','EOD_Level','EOD_Total_Bet_Placed','Number_of_Days_Played','Duration_of_Retention','EOD_$_Spent','dollar_spent'])
df.to_excel(path_export, sheet_name = "Day 0", index = False)

for i in [1,3,7]:
    book = load_workbook(path_export)
    writer = pd.ExcelWriter(path_export, engine = 'openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name= "Day {}".format(i), index = False)
    writer.save()
    writer.close()

# df_id = pd.read_excel("Dataset/7hc_ndroid_all-buyers.xlsx",sheet_name='Sheet1')
df_id = pd.read_csv("Dataset/7hc_andorid_philippines_Nov-10-Dec-10.csv")
# path_import =r'C:\Users\Dell\OneDrive\Python Scripts\PyCharm\Dataset\Session_History\check'
path_import = r'C:\Users\Dell\Downloads\Session_history\android'
# print(df_id.iloc[df_id.index[df_id['_id'] == '7HCAD2PKSUJW75B'],df_id.columns.get_loc('dollar_spent')])
# exit(0)

index = 0
excel_row = 1

df_day0 = pd.DataFrame()
df_day1 = pd.DataFrame()
df_day3 = pd.DataFrame()
df_day7 = pd.DataFrame()
df_collection = [df_day0,df_day1,df_day3,df_day7]

for id in df_id.iloc[:,df_id.columns.get_loc('_id')]:
    df = df_new()
    print('{}. {}'.format(excel_row, id))
    path_import_id = '{}\{}.csv'.format(path_import, id)

    try:
        df_import = pd.read_csv(path_import_id)
        # print(df_import.dtypes)
        if df_import.empty:
            print("No sesssion history")
            continue
        df_import.sort_values(by='created_at_ist',inplace = True, ascending=False)
    except OSError as e:
        print(e)
        continue

    list_datatype = ['coins', 'maximum_win', 'maximum_coins_reached', 'max_bet_placed','last_bet_placed', 'total_bet_amount','total_win_amount']
    df_import_datatype = pd.DataFrame(df_import[list_datatype])
    for col in df_import_datatype.columns:
        df_import_datatype[col] = df_import_datatype[col].astype('str')
        df_import_datatype[col] = df_import_datatype[col].str.replace(',','').astype('float')

    df_import[list_datatype] = df_import_datatype.loc[:, list_datatype]

    df_import['created_at_ist'] = pd.to_datetime(df_import['created_at_ist'])

    import_dates = np.array([])
    for row in range(len(df_import['created_at_ist'])):
        import_dates = np.append(import_dates,
                                     df_import.iloc[row, df_import.columns.get_loc('created_at_ist')].date())

    import_dates_unique = np.unique(import_dates)
    import_created = np.min(import_dates_unique)
    import_last_date = np.max(import_dates_unique)

    import_dates_final= np.array([])
    # for delta in [1,3,7]:
    #     import_dates_check = import_created + timedelta(days = delta)
    #     if import_dates_check in import_dates_unique:
    #         import_dates_final = np.append(import_dates_final,import_dates_check)
    import_dates_unique_sort = np.sort(import_dates_unique)
    # print("Total Played days = {}".format(import_dates_unique_sort.size))
# exit(0)
    for playdate in range(import_dates_unique_sort.size):
        day = (import_dates_unique_sort[playdate] - import_dates_unique_sort[0]).days
        if day == 0 or day == 1 or day == 3 or day == 7:
            import_dates_final = np.append(import_dates_final, import_dates_unique_sort[playdate])
        # import_dates_final = np.append(import_dates_final, import_dates_unique_sort[playdate])

    count = 0
    days = [0,1,3,7]
    for dates in np.sort(import_dates_final):
        print(dates)
        df.fillna(0, inplace=True)
        sheet_no = days[count]
        df_day = pd.DataFrame()
        for row in range(len(df_import)):
            if df_import.iloc[row, 0].date() == dates:
                df_day = df_day.append(df_import.iloc[row])
        # print(df_day)
        df_day = df_day[df_import.columns]
        df_day.sort_values(by="created_at_ist", inplace=True, ascending=False)

        df.loc[index, 'Daily_Total_Bet_Placed'] = (df_day.iloc[0, df_day.columns.get_loc('total_bet_amount')]) - (df_day.iloc[-1, df_day.columns.get_loc('total_bet_amount')])
        # print(df[index, 'Daily_Total_Bet_Placed'].item())
        df.loc[index, 'Daily_Max_Bet_Placed'] = df_day.loc[:, 'max_bet_placed'].max()
        df.loc[index, 'Daily_Games_Played'] = df_day.iloc[0, df_day.columns.get_loc('total_games_played')] - df_day.iloc[-1, df_day.columns.get_loc('total_games_played')]
        df.loc[index, 'Daily_Games_Win'] = df_day.iloc[0, df_day.columns.get_loc('total_games_win')] - df_day.iloc[-1, df_day.columns.get_loc('total_games_win')]
        df.loc[index, 'Daily_Games_Loose'] = df_day.iloc[0, df_day.columns.get_loc('total_games_loose')] - df_day.iloc[-1, df_day.columns.get_loc('total_games_loose')]
        df.loc[index, 'Daily_Level_Change'] = df_day.iloc[0, df_day.columns.get_loc('level')] - df_day.iloc[-1, df_day.columns.get_loc('level')]
        df.loc[index, 'Daily_Level_Up_Amount'] = df_day.iloc[0, df_day.columns.get_loc('total_level_up_amount')] - df_day.iloc[-1, df_day.columns.get_loc('total_level_up_amount')]
        df.loc[index, 'Daily_Win_Amount'] = df_day.iloc[0, df_day.columns.get_loc('total_win_amount')] - df_day.iloc[-1, df_day.columns.get_loc('total_win_amount')]
        df.loc[index, 'Daily_$_Spent'] = df_day.iloc[0, df_day.columns.get_loc('dollar_spent')] - df_day.iloc[-1, df_day.columns.get_loc('dollar_spent')]
        df.loc[index, 'EOD_Level_Up_Amount'] = df_day.iloc[0, df_day.columns.get_loc('total_level_up_amount')]
        df.loc[index, 'EOD_Win_Amount'] = df_day.iloc[0, df_day.columns.get_loc('total_win_amount')]
        df.loc[index, 'EOD_Coins'] = df_day.iloc[0, df_day.columns.get_loc('coins')]
        df.loc[index, 'EOD_Games_Played'] = df_day.iloc[0, df_day.columns.get_loc('total_games_played')]
        df.loc[index, 'EOD_Games_Win'] = df_day.iloc[0, df_day.columns.get_loc('total_games_win')]
        df.loc[index, 'EOD_Games_Loose'] = df_day.iloc[0, df_day.columns.get_loc('total_games_loose')]
        df.loc[index, 'EOD_Level'] = df_day.iloc[0, df_day.columns.get_loc('level')]
        df.loc[index, 'EOD_Total_Bet_Placed'] = df_day.iloc[0, df_day.columns.get_loc('total_bet_amount')]
        df.loc[index, 'Number_of_Days_Played'] = import_dates_unique.size
        df.loc[index, 'Duration_of_Retention'] = (import_last_date-import_created).days
        df.loc[index, 'EOD_$_Spent'] = df_day.iloc[0, df_day.columns.get_loc('dollar_spent')]
        df.loc[index, 'total_games_played'] = df_id.loc[df_id.index[df_id['_id'] == id],'total_games_played'].to_string(index = False)
        df.loc[index, 'dollar_spent'] = df_id.loc[df_id.index[df_id['_id'] == id],'dollar_spent'].to_string(index = False)
        df.loc[index, 'created_at_ist'] = df_import.iloc[-1, df_import.columns.get_loc('created_at_ist')]
        df.loc[index, '_id'] = id

        array_unique = df_day['machine_current_running_array'].unique()
        print(array_unique)
        for array in array_unique:
            df_array = pd.DataFrame()
            total_games_played = 0
            for row in range(len(df_day)):
                if df_day.iloc[row, df_day.columns.get_loc('machine_current_running_array')] == array:
                    df_array = df_array.append(df_day.iloc[row])

            df_array_index = df_day.index[df_day['machine_current_running_array'] == array].tolist()
            df_array_index.sort()
            if df_array.empty:
                continue

            df_array = df_array[df_import.columns]
            df_array.sort_values(by="created_at_ist", inplace=True, ascending=False)
            for i in range(len(df_array_index) - 1):
                if df_array_index[i + 1] - df_array_index[i] == 1:
                    total_games_played = total_games_played + (df_array.loc[df_array_index[i + 1], 'total_games_played'] - df_array.loc[df_array_index[i], 'total_games_played'])

            df.loc[index,array] = abs(total_games_played)
            df.fillna(0, inplace=True)
            for col in [300, 140, 97, 92, 85, '300.0', '140.0', '97.0', '92.0', '85.0']:
                df[col] = df[col].astype('float', errors = 'ignore')
            if index == 0:
                df.loc[index, str(array)] = df.iloc[index, df.columns.get_loc(str(array))] + df.iloc[index, df.columns.get_loc(array)]
            else:
                df.loc[index, str(array)] = float(df.iloc[index - 1, df.columns.get_loc(str(array))]) + float(df.iloc[index, df.columns.get_loc(array)])

        print("Day {}".format(sheet_no))
        print(df)

        # book = load_workbook(path_export)
        # writer = pd.ExcelWriter(path_export, engine='openpyxl')
        # writer.book = book
        # writer.sheets = {ws.title: ws for ws in book.worksheets}
        # workbook = writer.book
        # worksheet = writer.sheets["Day {}".format(sheet_no)]
        df_excel = pd.DataFrame(df.iloc[index:index + 1, :])

        if sheet_no == 0:
            df_day0 = df_day0.append(df_excel, ignore_index= True)
        elif sheet_no == 1:
            df_day1 = df_day1.append(df_excel, ignore_index= True)
        elif sheet_no == 3:
            df_day3 = df_day3.append(df_excel, ignore_index = True)
        elif sheet_no == 7:
            df_day7 = df_day7.append(df_excel, ignore_index = True)

        # df_excel.to_excel(writer,columns=df_excel.columns,sheet_name="Day {}".format(sheet_no), startrow=writer.sheets["Day {}".format(sheet_no)].max_row, header=False, index = False)
        # df_excel.to_excel(writer, sheet_name="Day {}".format(sheet_no),
        #                   startrow=writer.sheets["Day {}".format(sheet_no)].max_row, startcol=0, header=False,
        #                   index=False)
        # writer.save()
        # writer.close()
        index = index + 1
        count = count + 1

    del df
    excel_row = excel_row + 1
    index = 0

# print(df_day0)
# print(df_day1)
# print(df_day3)
# print(df_day7)
# days = [0,1,3,7]
# for i in range(len(df_collection)):

# for i in range(len(df_collection)):
book = load_workbook(path_export)
writer = pd.ExcelWriter(path_export, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
#
# df_collection[i].to_excel(writer, sheet_name="Day {}".format(days[i]),
#                           startrow=writer.sheets["Day {}".format(days[i])].max_row, startcol=0, header=False,
#                           index=False)

df_day0.to_excel(writer, sheet_name= 'Day 0',startrow=writer.sheets["Day 0"].max_row, startcol=0, header = False, index= False)
df_day1.to_excel(writer, sheet_name= 'Day 1',startrow=writer.sheets["Day 1"].max_row, startcol=0, header = False, index= False)
df_day3.to_excel(writer, sheet_name= 'Day 3',startrow=writer.sheets["Day 3"].max_row, startcol=0, header = False, index= False)
df_day7.to_excel(writer, sheet_name= 'Day 7',startrow=writer.sheets["Day 7"].max_row, startcol=0, header = False, index= False)

writer.save()
writer.close()

print("Execution time: {}".format(datetime.now()-begin_time))

